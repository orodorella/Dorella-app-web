from pathlib import Path

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
SOURCE = SCRIPT_DIR / "LIBRO GLOBAL REF.xlsx"
OUTPUT = SCRIPT_DIR / "04_product_supplier_data.sql"


def sql_value(value: object) -> str:
    if value is None or str(value).strip() == "":
        return "NULL"
    return "'" + str(value).strip().replace("'", "''") + "'"


workbook = load_workbook(SOURCE, read_only=True, data_only=True)
rows: list[tuple[str, object, object]] = []
seen_skus: set[str] = set()

for sheet in workbook.worksheets:
    for row in sheet.iter_rows(min_row=6, values_only=True):
        sku = row[3] if len(row) > 3 else None
        if sku is None or str(sku).strip() == "":
            continue
        normalized_sku = str(sku).strip()
        if normalized_sku in seen_skus:
            normalized_sku += "-DUP"
        seen_skus.add(str(sku).strip())
        proveedor = row[5] if len(row) > 5 else None
        referencia = row[6] if len(row) > 6 else None
        rows.append((normalized_sku, referencia, proveedor))

values = ",\n".join(
    f"  ({sql_value(sku)}, {sql_value(referencia)}, {sql_value(proveedor)})"
    for sku, referencia, proveedor in rows
)

OUTPUT.write_text(
    """-- Generado desde LIBRO GLOBAL REF.xlsx.
-- REF INTERNA -> products.sku
-- REFERENCIA -> products.referencia_proveedor
-- PROVEEDOR -> products.proveedor
WITH supplier_data (sku, referencia_proveedor, proveedor) AS (
VALUES
"""
    + values
    + """
)
UPDATE products AS p
SET referencia_proveedor = d.referencia_proveedor,
    proveedor = d.proveedor,
    material = NULL,
    updated_at = now()
FROM supplier_data AS d
WHERE p.sku = d.sku;
""",
    encoding="utf-8",
)

print(f"Generadas {len(rows)} filas en {OUTPUT}")
